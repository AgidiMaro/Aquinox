# Import necessary modules
from PyPDF2 import PdfReader
import csv
import docx
import pythoncom
from win32com import client as win32
from openpyxl import load_workbook

# Class to handle file reading operations
class FileReader:
    
    # Method to get the file path for a given file name
    def get_file_path(self, file_name):
        result = 'audit-templates/' + file_name + '.csv'  # Construct the file path
        return result

    # Method to read a PDF file and extract its text
    def read_pdf_file(self, file) -> str:
        text = ""
        pdf_reader = PdfReader(file)  # Initialize the PDF reader with the file
        
        # Extract text from each page of the PDF
        for page in pdf_reader.pages:
            text += page.extract_text()

        return text
    
    # Method to read a text file and extract its content
    def read_txt_file(self, file) -> str:
        text = file.read().decode('utf-8')  # Read the text file content and decode it
        return text
    
    # Method to read a DOCX file and extract its content
    def read_docx_file(self, file) -> str:
        doc = docx.Document(file)
        text = '\n'.join([paragraph.text for paragraph in doc.paragraphs])
        return text

    # Method to read a DOC file and extract its content
    def read_doc_file(self, file_path) -> str:
        pythoncom.CoInitialize()
        word = win32.Dispatch("Word.Application")
        doc = word.Documents.Open(file_path)
        text = doc.Range().Text
        doc.Close()
        word.Quit()
        return text
    
    # Method to read an XLSX file and extract its content
    def read_xlsx_file(self, file) -> str:
        workbook = load_workbook(file)
        text = ""
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                text += row_text + "\n"
        return text
    
    # Method to read a CSV file and extract its content as plain text
    def read_csv_file(self, file) -> str:
        text = ""
        file.stream.seek(0)  # Ensure the file pointer is at the beginning
        csv_reader = csv.reader(file.stream.read().decode('utf-8').splitlines())
        for row in csv_reader:
            text += ", ".join(row) + "\n"
        return text

    # Method to read a CSV D&I template file and return its contents as a dictionary
    def read_csv_template_file(self, file_path):
        file_result = {}  # Dictionary to store the results
        domain = None
        questions_list = []

        current_domain = None
        current_question = None
        current_example = None
        

        # Open and read the CSV file
        with open(file_path) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0

            # Iterate through each row in the CSV file
            for row in csv_reader:
                line_count += 1
                
                # Skip the header row
                if line_count == 1:
                    continue

                current_domain = row[0].strip()  # Read the domain from the first column
                current_question = row[1].strip()  # Read the question from the second column

                # If the current domain is empty, use the previous domain
                if not current_domain:
                    current_domain = domain
                else:
                    domain = current_domain
                    questions_list = []
                    file_result[domain] = questions_list  # Initialize a new list for the new domain

                questions_list.append({"question": current_question, "example": current_example})  # Add the question and example to the list

            print(f'Processed {line_count} lines.')

        return file_result
